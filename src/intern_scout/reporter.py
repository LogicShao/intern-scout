import json
from intern_scout.models import Position, SearchResult


def to_json(result: SearchResult, indent: int = 2) -> str:
    return json.dumps(result.to_dict(), ensure_ascii=False, indent=indent)


def to_markdown_table(result: SearchResult, max_rows: int = 50) -> str:
    if not result.positions:
        return f"_(empty)_ — {result.source}: {result.message}"

    lines: list[str] = []
    lines.append(f"## {result.source} ({result.fetched}/{result.total} positions)\n")
    lines.append("| # | 岗位 | 部门 | 城市 | 薪资 |")
    lines.append("|---|------|------|------|------|")

    for i, p in enumerate(result.positions[:max_rows], 1):
        title = p.title[:40]
        dept = p.department[:20]
        loc = p.location[:15]
        salary = p.salary_range[:15] or "-"
        lines.append(f"| {i} | [{title}]({p.source_url}) | {dept} | {loc} | {salary} |")

    if result.fetched > max_rows:
        lines.append(f"\n_(showing first {max_rows} of {result.fetched} positions)_")
    return "\n".join(lines)


def to_excel(result: SearchResult, filepath: str):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = result.source[:31]

    headers = ["post_id", "title", "company", "department", "location",
               "salary_range", "recruit_type", "posted_date", "source_url"]
    ws.append(headers)

    for p in result.positions:
        ws.append([
            p.post_id, p.title, p.company, p.department, p.location,
            p.salary_range, p.recruit_type, p.posted_date, p.source_url,
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(filepath)


def format_output(results: list[SearchResult], fmt: str = "json", filepath: str = ""):
    if fmt == "json":
        combined = {"results": [r.to_dict() for r in results]}
        output = json.dumps(combined, ensure_ascii=False, indent=2)
        if filepath:
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(output)
        return output

    if fmt == "table":
        return "\n\n".join(to_markdown_table(r) for r in results)

    if fmt == "excel":
        path = filepath or "output/intern_positions.xlsx"
        import os
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        if len(results) == 1:
            to_excel(results[0], path)
        else:
            import openpyxl
            wb = openpyxl.Workbook()
            for i, r in enumerate(results):
                ws = wb.create_sheet(title=r.source[:31]) if i > 0 else wb.active
                if ws is None:
                    ws = wb.create_sheet(title=f"sheet_{i}")
                headers = ["post_id", "title", "company", "department", "location",
                           "salary_range", "recruit_type", "posted_date", "source_url"]
                ws.append(headers)
                for p in r.positions:
                    ws.append([
                        p.post_id, p.title, p.company, p.department, p.location,
                        p.salary_range, p.recruit_type, p.posted_date, p.source_url,
                    ])
                for col in ws.columns:
                    max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
            default_sheet = wb.get_sheet_by_name("Sheet")
            if default_sheet and len(results) > 1:
                wb.remove(default_sheet)
            wb.save(path)
        return f"Saved to {path}"

    return to_json(results[0]) if results else ""
